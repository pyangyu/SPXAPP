import tkinter as tk
from tkinter import filedialog, messagebox

import openpyxl
from PIL import Image, ImageTk
import re
import os
import pandas as pd
from tkinter import messagebox
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

pattern1 = r"\d{3}-\d{8}"
pattern2 = r"\d{11}"


def extract_file_name(file_name):
    file_name = file_name.replace(" ", "")
    matches = re.findall(pattern1, file_name)
    if not matches:
        matches = re.findall(pattern2, file_name)
        if matches:
            matches = matches[0]
            return matches[:3] + "-" + matches[3:]
        else:
            return []
    if len(matches) > 1:
        return ""
    return matches[0]


def combine_files_new():
    # get the selected folder path
    folder_path = folder_path_var.get()

    # create a new subdirectory called "complete_data_pool" under the specified folder path
    complete_data_pool_files = os.path.join(folder_path, "completed data pool")
    os.makedirs(complete_data_pool_files, exist_ok=True)

    driver_info_path = os.path.join(folder_path, "driver info")
    driver_info_file = [f for f in os.listdir(driver_info_path) if f.endswith(".xlsx")]
    if len(driver_info_file) > 1:
        messagebox.showinfo("Combine Files", "Something is wrong, please contact me")
        return
    driver_info_file = driver_info_file[0]
    dataframe_driver_info = pd.read_excel(os.path.join(driver_info_path, driver_info_file), sheet_name="Merged tables")
    dataframe_driver_info = dataframe_driver_info[["Route", "Tracking Number"]]

    # get a list of all Excel files in the folder
    excel_files = [f for f in os.listdir(folder_path) if f.endswith(".xlsx")]

    # Specify the desired order of columns
    desired_order = ['Tracking Number', 'MAWB', 'Courier code', 'Action', 'DSP', 'consignor_item_id', 'display_id',
                     'receptacle_id', 'sender_name', 'sender_orgname', 'sender_address1', 'sender_address2',
                     'sender_district', 'sender_city', 'sender_state', 'sender_zip5', 'sender_zip4', 'sender_country',
                     'sender_phone', 'sender_email', 'sender_url', 'recipient_name', 'recipient_orgname',
                     'recipient_address1', 'recipient_address2', 'recipient_district', 'recipient_city',
                     'recipient_state', 'recipient_zip5', 'recipient_zip4', 'recipient_country', 'recipient_phone',
                     'recipient_email', 'recipient_addr_type', 'return_name', 'return_orgname', 'return_address1',
                     'return_address2', 'return_district', 'return_city', 'return_state', 'return_zip5', 'return_zip4',
                     'return_country', 'return_phone', 'return_email', 'mail_type', 'pieces', 'weight', 'length',
                     'width', 'height', 'girth', 'value', 'machinable', 'po_box_flag', 'gift_flag', 'commercial_flag',
                     'customs_quantity_units', 'dutiable', 'duty_pay_by', 'product', 'description', 'url', 'sku',
                     'country_of_origin', 'manufacturer', 'harmonization_code', 'unit_value', 'quantity', 'total_value',
                     'total_weight']

    # desired_order = pd.read_excel(os.path.join(folder_path, "0520 - Data pool.xlsx"), sheet_name='Edit').columns.tolist()
    # messagebox.showinfo("Combine Files", str(desired_order))
    # print(desired_order)
    # return

    # iterate through each Excel file and concatenate all columns
    for file_name in excel_files:
        # the extract file_name will be the MAWB column
        MAWB_number = extract_file_name(file_name)
        if MAWB_number == "":
            messagebox.showinfo("Combine Files", "Something is wrong, please contact me")
            return

        # go into the file
        file_path = os.path.join(folder_path, file_name)
        # read the Excel file into a Pandas dataframe
        dataframe_T86 = pd.read_excel(os.path.join(folder_path, file_path))

        # create a new file name data_pool.xlsx, and store in under the file path
        new_file_name = MAWB_number + "_for_data_pool.xlsx"
        new_file_path = os.path.join(complete_data_pool_files, new_file_name)

        # concatenate the columns
        MAWB_column = pd.DataFrame({"MAWB": [MAWB_number] * len(dataframe_T86)})
        new_dataframe_T86 = pd.concat([dataframe_T86, MAWB_column], axis=1)
        # Create a new DataFrame with an empty "Action" column
        action_column = pd.DataFrame({"Action": [None] * len(dataframe_T86)})
        new_dataframe_T86 = pd.concat([new_dataframe_T86, action_column], axis=1)
        # Create the new "status" column based on the condition
        DSP_column = dataframe_T86["consignor_item_id"].apply(
            lambda x: "UD" if "UD" in x else "AH" if "AH" in x else "GE" if "GE" in x else "**wrong**")
        DSP_column = pd.DataFrame({"DSP": DSP_column})
        new_dataframe_T86 = pd.concat([new_dataframe_T86, DSP_column], axis=1)
        # Create the new "Tracking Number" column the same as "consignor_item_id" and concat
        Tracking_Number = pd.DataFrame({"Tracking Number": dataframe_T86['consignor_item_id']})
        new_dataframe_T86 = pd.concat([new_dataframe_T86, Tracking_Number], axis=1)
        # Create the new "Courtier Code" column the same as "consignor_item_id" and concat
        DSP_backup_column = dataframe_T86["consignor_item_id"].apply(
            lambda x: "SPXUD" if "UD" in x else "SPXAH" if "AH" in x else "SPXGE" if "GE" in x else "**wrong**")
        DSP_backup_column = pd.DataFrame({"Courier code": DSP_backup_column})
        new_dataframe_T86 = pd.concat([new_dataframe_T86, DSP_backup_column], axis=1)

        # Create the new "Routes" column based on the "Tracking Number"
        new_dataframe_T86['Route'] = ''
        for index, row in new_dataframe_T86.iterrows():
            if row['Tracking Number'] in dataframe_driver_info['Tracking Number'].values:
                new_dataframe_T86.at[index, 'Route'] = dataframe_driver_info.loc[
                    dataframe_driver_info['Tracking Number'] == row['Tracking Number'], 'Route'].values[0]

        # rename the column name
        if "harmonization_code" in new_dataframe_T86.columns:
            new_dataframe_T86.rename(columns={"harmonization_code": "Harmonization_code"}, inplace=True)

        if 'recipient_zip5' in new_dataframe_T86.columns and 'recipient_zip4' in new_dataframe_T86.columns:
            new_dataframe_T86['recipient_zip4'] = new_dataframe_T86.apply(
                lambda row: row['recipient_zip5'][-4:] if len(str(row['recipient_zip5'])) > 5 and
                                                          row['recipient_zip4'] == "" else row['recipient_zip4'] if row[
                                                                                                                        'recipient_zip4'] != "" else "",
                axis=1)
            # Update the values in the 'zip code' column using a lambda function
            new_dataframe_T86['recipient_zip5'] = new_dataframe_T86['recipient_zip5'].apply(
                lambda x: str(x).split("-")[0] if len(str(x).split("-")) > 1 else str(x))

        # rank the columns
        # Get the remaining columns
        for column in desired_order:
            if column not in new_dataframe_T86.columns:
                new_dataframe_T86[column] = ''
        new_dataframe_T86 = new_dataframe_T86[
            desired_order + new_dataframe_T86.columns.difference(desired_order).tolist()]
        desired_order = desired_order

        ''' creating at the same time '''
        '''
        # Apply lambda function to create "DSP" and "DSP_backup" columns
        columns = dataframe_T86["consignor_item_id"].apply(lambda x: (
            "UD" if "UD" in x else "AH" if "AH" in x else "GE" if "GE" in x else "**wrong**",
            "SPXUD" if "UD" in x else "SPXAH" if "AH" in x else "SPXGE" if "GE" in x else "**wrong**"
        ))

        # Create DataFrame for "DSP" and "DSP_backup" columns
        columns = pd.DataFrame(columns.tolist(), columns=["DSP", "DSP_backup"])

        # Concatenate the new columns with the existing DataFrame
        new_dataframe_T86 = pd.concat([new_dataframe_T86, columns], axis=1)
        '''

        # create a new workbook and worksheet using openpyxl
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # write the dataframe to the worksheet
        for row in dataframe_to_rows(new_dataframe_T86, index=False, header=True):
            worksheet.append(row)

        # save the workbook to the new Excel file
        workbook.save(new_file_path)

    dataframes = []

    # Iterate over each file in the folder
    for file in os.listdir(complete_data_pool_files):
        if file.endswith('.xlsx'):
            file_path = os.path.join(complete_data_pool_files, file)
            dataframe = pd.read_excel(file_path)
            dataframes.append(dataframe)

    # Concatenate the dataframes into a single dataframe
    combined_dataframe = pd.concat(dataframes, ignore_index=True)

    # Reorder the columns based on the first file's column order
    combined_dataframe = combined_dataframe[dataframes[0].columns]

    # create a new workbook and worksheet using openpyxl
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Define the columns to be highlighted
    columns_to_highlight = ['Tracking Number', 'MAWB', 'Courier code', 'Action', 'DSP']

    # Write the dataframe to the worksheet
    for r in dataframe_to_rows(combined_dataframe, index=False, header=True):
        worksheet.append(r)

    # Iterate over the columns and update the header style
    for col_num, col_value in enumerate(worksheet.iter_cols(min_row=1, max_row=1), 1):
        # print(col_value)
        # print(type(col_value))
        # print(col_value[0])
        # print(type(col_value[0]))
        if col_value[0].value in columns_to_highlight:
            # Set the font color to red
            col_value[0].font = openpyxl.styles.Font(color="FF0000")

    # Iterate over the columns and update the font color for the values
    for col_num, col_value in enumerate(worksheet.iter_cols(min_row=2), 1):
        if combined_dataframe.columns[col_num - 1] in columns_to_highlight:
            # Set the font color to light blue for each cell in the column
            for cell in col_value:
                cell.font = openpyxl.styles.Font(color='4796CB')

    workbook.save(os.path.join(folder_path, "combine_completed.xlsx"))

    # if not df_list:
    #     # if there are no files with "Tracking Number" column, show a message box and return
    #     messagebox.showinfo("Combine Files", "No files with 'Tracking Number' column found.")
    #     return

    # # concatenate all the DataFrames together
    # combined_df = pd.concat(df_list, ignore_index=True)
    #
    # # write the combined DataFrame to a new Excel file
    # combined_file_path = os.path.join(folder_path, "combined.xlsx")
    # combined_df.to_excel(combined_file_path, index=False)

    # display a message box to indicate the operation is complete
    messagebox.showinfo("Combine Files", "The files have been successfully combined.")


def browse_folder():
    folder_path = filedialog.askdirectory()
    folder_path_var.set(folder_path)


if __name__ == "__main__":
    # initialize the UI
    root = tk.Tk()
    root.title("Filter T86")

    # set the window size
    root.geometry("500x500")

    # create a canvas widget
    canvas = tk.Canvas(root, width=500, height=500)
    # create an image object from the icon.webp file
    # get the selected folder path
    image_folder_path = os.path.join(os.getcwd(), 'combine_icon.png')
    image = Image.open(image_folder_path)
    # adjust the alpha channel to 0.3
    image.putalpha(int(255 * 0.2))
    photo_image = ImageTk.PhotoImage(image)
    # set the canvas background color to #DAE6E6
    canvas.configure(bg='#DAE6E6')
    # create a rectangle with the same size as the canvas to serve as the background
    background = canvas.create_rectangle(0, 0, 500, 500, fill="#DAE6E6", outline="#DAE6E6")
    # create an image item on the canvas with the icon.webp image
    canvas.create_image(0, 0, image=photo_image, anchor="nw")
    # pack the canvas widget to fill the window
    canvas.pack(fill="both", expand=True)

    # add a label and entry for the folder path
    folder_path_var = tk.StringVar()
    folder_path_label = tk.Label(root, text="Folder Path:", font=("Helvetica", 20, "bold"), fg="darkblue", bg="#DAE6E6")
    folder_path_label.pack(side=tk.TOP)
    folder_path_label.place(relx=0.5, rely=0.2, anchor=tk.CENTER)
    folder_path_entry = tk.Entry(root, textvariable=folder_path_var, width=40, font=("Helvetica", 14))
    folder_path_entry.pack(side=tk.TOP)
    folder_path_entry.place(relx=0.5, rely=0.4, anchor=tk.CENTER)

    browse_button = tk.Button(root, text="Browse", command=browse_folder, font=("Helvetica", 12), bg="orange",
                              bd=2, relief=tk.RAISED, activebackground="#FF9999", activeforeground="white",
                              padx=10, pady=5)
    browse_button.pack(pady=10, side=tk.TOP)
    browse_button.place(relx=0.5, rely=0.6, anchor=tk.CENTER)

    # add a button to perform the combination
    combine_button = tk.Button(root, text="Filter T86 Files", command=combine_files_new, font=("Helvetica", 12),
                               bg="lightblue",
                               bd=2, relief=tk.RAISED, activebackground="#FF9999", activeforeground="white",
                               padx=10, pady=5)
    combine_button.pack(pady=5, side=tk.BOTTOM)
    combine_button.place(relx=0.5, rely=0.8, anchor=tk.CENTER)

    # start the UI loop
    root.mainloop()
